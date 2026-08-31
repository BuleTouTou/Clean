from __future__ import annotations

import csv, hashlib, io, json, os, re, shutil, subprocess, threading, time, unicodedata, uuid, webbrowser
from copy import copy
from datetime import date, datetime, timezone
from difflib import SequenceMatcher
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse, unquote

import openpyxl
from .db import init_db, load_setting, save_report, save_setting
from .storage import get_storage

# core.py 位于 backend/ 包内，项目根目录是其父目录。
ROOT = Path(__file__).resolve().parent.parent
DATA = Path(os.getenv("WORK_DIR", str(ROOT / "data"))).resolve()
UPLOADS = DATA / "uploads"
OUTPUTS = DATA / "outputs"
RESOURCES = ROOT / "resources"
RULES_FILE = ROOT / "data" / "rules.json"
DOWNLOADS_FILE = DATA / "downloads.json"
STATIC = ROOT / "frontend" / "dist"
for p in (DATA, UPLOADS, OUTPUTS, RESOURCES): p.mkdir(parents=True, exist_ok=True)
init_db()


def utc_now_iso() -> str:
    """Return a compact, timezone-aware UTC ISO 8601 timestamp."""
    return datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")

FILES = {
    "estate": "北京楼盘字典.xlsx",
    "unit": "北京-单元楼盘字典.xlsx",
    "sale": "售房数据导入模版.xlsx",
    "rent": "租房数据导入模版 .xlsx",
}
RULE_VERSION = "1.0.0"
TASKS: dict[str, dict] = {}
DOWNLOADS: dict[str, Path] = {}
LOCK = threading.RLock()

SYNONYMS = {
    "城区": ["行政区", "城区", "区域", "所属城区"], "商圈": ["片区", "商圈", "板块"],
    "小区": ["小区", "楼盘", "项目名称", "社区"], "栋幢": ["座栋", "栋幢", "楼栋", "栋号"],
    "房号": ["房号", "门牌号", "室号"], "面积": ["建筑面积", "建面", "面积", "面积㎡"],
    "总价(万)": ["售价", "出售价格", "总价"], "月租金": ["租价", "租金", "月租", "出租价格"],
    "业主姓名": ["业主", "房东", "产权人"], "业主联系方式": ["电话", "手机号", "联系方式", "业主电话"],
    "录入人": ["录入人姓名", "录入员工", "录入人员"], "维护人": ["维护人姓名", "维护员工", "维护人员"],
    "等级": ["等级", "房源等级"],
}
# 工号和部门代码不再无条件清空：只有用户明确映射且源值非空时才写入。
# 未映射字段仍保持模板空值，系统不会推测或生成任何代码。
SYSTEM_BLANK = re.compile(r"(?!)")
TEXT_FIELD = re.compile(r"(编号|电话|联系方式|工号|代码)$")

def jload(path: Path, default):
    if path == RULES_FILE:
        try:
            legacy_default = json.loads(path.read_text("utf-8")) if path.is_file() else default
        except Exception:
            legacy_default = default
        return load_setting("cleaning_rules", legacy_default)
    try: return json.loads(path.read_text("utf-8"))
    except Exception: return default

def jsave(path: Path, obj):
    if path == RULES_FILE:
        save_setting("cleaning_rules", obj)
        return
    path.write_text(json.dumps(obj, ensure_ascii=False, indent=2), "utf-8")

def is_managed_path(path: Path) -> bool:
    resolved = path.resolve()
    return any(root.resolve() == resolved or root.resolve() in resolved.parents for root in (ROOT, DATA))

def register_download(path: Path):
    token=uuid.uuid4().hex
    registry=jload(DOWNLOADS_FILE,{})
    registry[token]=str(path.resolve())
    jsave(DOWNLOADS_FILE,registry)
    DOWNLOADS[token]=path.resolve()
    return token

def resolve_download(token):
    fp=DOWNLOADS.get(token)
    if fp: return fp
    raw=jload(DOWNLOADS_FILE,{}).get(token)
    if raw:
        fp=Path(raw).resolve()
        if is_managed_path(fp):
            DOWNLOADS[token]=fp
            return fp
    return None

def get_task(task_id):
    task=TASKS.get(task_id)
    if task is None:
        raise ValueError("当前批次已因工具重启而失效，请返回首页重新创建批次")
    return task

def norm(v, compact=False):
    s = unicodedata.normalize("NFKC", str(v or "")).strip().replace("\r", "").replace("\n", "")
    s = s.casefold()
    if compact:
        s = re.sub(r"[（(][^）)]*[）)]", "", s)
        s = re.sub(r"[\s·•,，.。\-_—/\\#号座栋幢楼]", "", s)
    return s

def aux_header(v): return re.sub(r"[（(][^）)]*[）)]", "", norm(v)).strip()

def clean_import_value(value):
    if isinstance(value,str): return re.sub(r"^'+\s*","",value.strip())
    return value

def parse_layout(value):
    text=unicodedata.normalize("NFKC",str(value or "")).strip()
    if not text: return {}
    cn={"零":0,"〇":0,"一":1,"二":2,"两":2,"三":3,"四":4,"五":5,"六":6,"七":7,"八":8,"九":9,"十":10}
    result={}
    for amount,label in re.findall(r"(\d+|[零〇一二两三四五六七八九十]+)\s*(房|室|厅|厨|卫|阳台)",text):
        if amount.isdigit(): number=int(amount)
        elif amount in cn: number=cn[amount]
        elif amount.startswith("十"): number=10+cn.get(amount[1:],0)
        elif amount.endswith("十"): number=cn.get(amount[:-1],0)*10
        else: number=cn.get(amount[0],0)*10+cn.get(amount[1:],0)
        field="室" if label=="房" else label
        result[field]=number
    return result

def file_status():
    return [{"key": k, "name": n, "exists": (RESOURCES/n).exists(), "size": (RESOURCES/n).stat().st_size if (RESOURCES/n).exists() else 0,
             "mtime": (RESOURCES/n).stat().st_mtime if (RESOURCES/n).exists() else None} for k,n in FILES.items()]

def template_info(kind):
    p = RESOURCES / FILES[kind]
    wb = openpyxl.load_workbook(p, read_only=True, data_only=False)
    if "房源" not in wb.sheetnames: raise ValueError(f"模板 {p.name} 缺少‘房源’工作表")
    ws = wb["房源"]
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    if not all(headers): raise ValueError("模板表头包含空列")
    return {"path": p, "headers": headers, "signature": hashlib.sha256(("|".join(headers)+str(p.stat().st_mtime_ns)).encode()).hexdigest()}

def detect_header(rows):
    best = (0, -1)
    for i,row in enumerate(rows[:30]):
        vals = [str(x).strip() for x in row if x not in (None, "")]
        unique = len(set(vals)); textish = sum(any(ch.isalpha() or '\u4e00' <= ch <= '\u9fff' for ch in x) for x in vals)
        score = unique * 2 + textish - max(0, len(vals)-unique)*3
        if len(vals) >= 2 and score > best[1]: best = (i, score)
    return best[0]

def csv_rows(path):
    raw = path.read_bytes(); encoding = None
    for enc in ("utf-8-sig", "utf-8", "gb18030", "gbk"):
        try: text = raw.decode(enc); encoding = enc; break
        except UnicodeDecodeError: pass
    if encoding is None: raise ValueError("无法识别 CSV 编码")
    sample = text[:65536]
    try: dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|，")
    except csv.Error: dialect = csv.excel
    rows = list(csv.reader(io.StringIO(text), dialect))
    return rows, encoding, dialect.delimiter

def convert_xls(path):
    target = path.with_suffix(".xlsx")
    soffice = shutil.which("soffice")
    if soffice:
        subprocess.run([soffice,"--headless","--convert-to","xlsx","--outdir",str(path.parent),str(path)], check=True, timeout=120)
        return target
    raise ValueError("旧版 XLS 需要本机安装 LibreOffice；请另存为 XLSX 后重试")

def inspect_upload(path):
    ext = path.suffix.lower(); sheets=[]
    if ext == ".csv":
        rows, enc, delim = csv_rows(path); hi=detect_header(rows); width=max((len(r) for r in rows), default=0)
        sheets.append({"name":"CSV", "hidden":False, "rows":max(0,len(rows)-hi-1), "cols":width, "headerRow":hi+1,
                       "preview":[list(map(str,r[:20])) for r in rows[:5]], "encoding":enc, "delimiter":delim})
    else:
        if ext == ".xls": path=convert_xls(path)
        wb=openpyxl.load_workbook(path, read_only=False, data_only=True)
        for ws in wb.worksheets:
            sample=list(ws.iter_rows(min_row=1,max_row=min(ws.max_row,30),values_only=True)); hi=detect_header(sample)
            value_rows=[cell.row for cell in ws._cells.values() if cell.value not in (None,"")]
            last_value_row=max(value_rows,default=hi+1)
            sheets.append({"name":ws.title,"hidden":ws.sheet_state!="visible","rows":max(0,last_value_row-hi-1),"cols":ws.max_column,
                           "headerRow":hi+1,"preview":[["" if v is None else str(v) for v in r[:20]] for r in sample[:5]],
                           "merged":[str(x) for x in list(ws.merged_cells.ranges)[:50]]})
    return path, sheets

def read_table(path, sheet, header_row):
    if path.suffix.lower()==".csv":
        all_rows,_,_=csv_rows(path); row_iter=iter(all_rows)
    else:
        wb=openpyxl.load_workbook(path, read_only=True, data_only=True); ws=wb[sheet]; row_iter=ws.iter_rows(values_only=True)
    for _ in range(header_row-1): next(row_iter,None)
    header_values=next(row_iter,())
    raw=["" if x is None else str(x).strip() for x in header_values]
    while raw and raw[-1]=="": raw.pop()
    seen={}; headers=[]
    for i,h in enumerate(raw):
        h=h or f"未命名列{i+1}"; seen[h]=seen.get(h,0)+1; headers.append(h if seen[h]==1 else f"{h}_{seen[h]}")
    data=[]; blank_run=0; seen_data=False
    for rn,row in enumerate(row_iter, start=header_row+1):
        vals=[clean_import_value(v) for v in row[:len(headers)]]+[None]*max(0,len(headers)-len(row))
        if any(v not in (None,"") for v in vals):
            data.append({"_row":rn, **dict(zip(headers,vals))}); blank_run=0; seen_data=True
        elif seen_data:
            blank_run+=1
            if blank_run>=500: break
    return headers,data

def mapping_suggestions(headers, target, rows, source_key):
    rules=jload(RULES_FILE, {"mappings":{},"estates":{},"buildings":{},"units":{}})
    saved=rules.get("mappings",{}).get(source_key,{})
    syn={norm(x):k for k,vals in SYNONYMS.items() for x in vals}
    out=[]
    for h in headers:
        nh=norm(h); ah=aux_header(h); suggestion=""; confidence=0; reason="需人工选择"
        if h in target: suggestion=h; confidence=1; reason="字段名完全一致"
        elif nh in syn and syn[nh] in target: suggestion=syn[nh]; confidence=.98; reason="同义词完全匹配"
        elif h in saved and saved[h] in target: suggestion=saved[h]; confidence=.97; reason="已确认来源映射"
        elif ah in [norm(x) for x in target]: suggestion=target[[norm(x) for x in target].index(ah)]; confidence=.9; reason="去除单位后匹配"
        else:
            scored=sorted(((SequenceMatcher(None,ah,norm(x)).ratio(),x) for x in target),reverse=True)
            if scored and scored[0][0]>=.55: confidence,suggestion=scored[0]; reason="名称相似，仅供建议"
        samples=[]
        for r in rows:
            if r.get(h) not in (None,""):
                samples.append(str(r.get(h)))
                if len(samples)>=5: break
        out.append({"source":h,"target":suggestion,"confidence":round(confidence,2),"reason":reason,"samples":samples})
    return out

def load_estates():
    wb=openpyxl.load_workbook(RESOURCES/FILES["estate"],read_only=True,data_only=True); ws=wb[wb.sheetnames[0]]
    rows=ws.iter_rows(values_only=True); heads=[str(x or "").strip() for x in next(rows)]; ix={h:i for i,h in enumerate(heads)}
    result=[]
    for r in rows:
        if not r[ix["楼盘"]]: continue
        result.append({k:("" if r[ix[k]] is None else str(r[ix[k]]).strip()) for k in ("楼盘","别名","行政区","商圈","期名")})
    return result

def estate_output_name(estate):
    """模板小区名使用楼盘+期名；楼盘名中的（一区）/（一期）保留文字但去掉括号。"""
    base=str(estate.get("楼盘") or "").strip()
    phase=str(estate.get("期名") or "").strip()
    base=re.sub(r"[（(]\s*((?:第?[一二三四五六七八九十百零〇两\d]+期)|(?:[一二三四五六七八九十百零〇两\d]+区))\s*[）)]",r"\1",base)
    if not phase or norm(base,True).endswith(norm(phase,True)): return base
    return base+phase

def estate_match_names(estate):
    names={str(estate.get("楼盘") or "").strip(),estate_output_name(estate)}
    phase=str(estate.get("期名") or "").strip()
    for alias_name in re.split(r"[,，、;/；|]",str(estate.get("别名") or "")):
        alias_name=alias_name.strip()
        if not alias_name: continue
        names.add(alias_name)
        if phase and not norm(alias_name,True).endswith(norm(phase,True)): names.add(alias_name+phase)
    return {x for x in names if x}

def estate_norm(value, compact=False):
    # 小区名里的（一区）/（一期）参与匹配，只忽略括号本身。
    text=re.sub(r"[（）()]","",str(value or ""))
    zone_map={"一":"A","二":"B","两":"B","三":"C","四":"D","五":"E","六":"F","七":"G","八":"H","九":"I","十":"J"}
    def zone_repl(match):
        token=match.group(1).upper()
        if token.isdigit() and 1<=int(token)<=26: token=chr(64+int(token))
        else: token=zone_map.get(token,token)
        return token+"区"
    text=re.sub(r"(?<![A-Za-z0-9一二两三四五六七八九十])([A-Za-z]|\d{1,2}|[一二两三四五六七八九十])区",zone_repl,text,flags=re.I)
    return norm(text,compact)

def phase_hint(value):
    text=unicodedata.normalize("NFKC",str(value or ""))
    match=re.search(r"(?<![A-Za-z0-9一二两三四五六七八九十])([A-Za-z]|\d{1,2}|[一二两三四五六七八九十])区",text,re.I)
    if not match: return ""
    return estate_norm(match.group(0),False).upper()

def community_decision_key(raw_community, raw_building=""):
    hint=phase_hint(raw_community) or phase_hint(raw_building)
    raw=str(raw_community or "").strip(); building=str(raw_building or "").strip()
    if building: raw+="｜原始栋座："+building
    if hint: raw+="｜期名提示："+hint
    return raw

def unique_estates(items):
    unique={}
    for estate in items:
        key=(estate_output_name(estate),estate.get("行政区",""),estate.get("商圈",""))
        unique.setdefault(key,estate)
    return list(unique.values())

def building_courtyard_prefixes(buildings):
    prefixes=set()
    for building in buildings:
        m=re.match(r"^(.+?号院)",str(building or "").strip())
        if m: prefixes.add(m.group(1))
    return prefixes

def community_building_prefix(raw_community, estate, unit_index):
    """仅当小区尾部的“几号院”确实出现在该楼盘栋座字典中时才拆出。"""
    buildings=unit_index.get(estate.get("楼盘",""),{})
    for prefix in sorted(building_courtyard_prefixes(buildings),key=len,reverse=True):
        for estate_name in estate_match_names(estate):
            if estate_norm(estate_name+prefix,True)==estate_norm(raw_community,True): return prefix
    return ""

def merge_building_prefix(prefix, raw_building):
    prefix=str(prefix or "").strip(); raw_building=str(raw_building or "").strip()
    if not prefix: return raw_building
    if not raw_building: return prefix
    if estate_norm(raw_building,True).startswith(estate_norm(prefix,True)): return raw_building
    return prefix+raw_building

def building_decision_key(estate, raw_building):
    return estate_output_name(estate)+"|"+str(raw_building or "").strip()

def estate_review(task):
    mapping=task["mapping"]; rows=task["rows"]; estates=load_estates(); unit_index=load_unit_index(); rules=jload(RULES_FILE,{})
    src_by_target={}
    for src,tgts in mapping.items():
        for t in (tgts if isinstance(tgts,list) else [tgts]):
            if t and t!="__ignore__": src_by_target.setdefault(t,[]).append(src)
    community_src=src_by_target.get("小区",[])
    building_src=src_by_target.get("栋幢",[])
    if not community_src: return [], [{"type":"小区没有映射字段","count":len(rows)}]
    exact={}; compact={}
    for e in estates:
        e["match_name"]=estate_output_name(e)
        names=set(estate_match_names(e))
        for prefix in building_courtyard_prefixes(unit_index.get(e["楼盘"],{})):
            names.update(name+prefix for name in estate_match_names(e))
        for name in names:
            exact.setdefault(estate_norm(name),[]).append(e); compact.setdefault(estate_norm(name,True),[]).append(e)
    groups={}
    for r in rows:
        raw=" ".join(str(r.get(s) or "").strip() for s in community_src).strip()
        raw_build=" ".join(str(r.get(s) or "").strip() for s in building_src).strip()
        key=community_decision_key(raw,raw_build); hint=phase_hint(raw) or phase_hint(raw_build)
        group=groups.setdefault(key,{"raw":raw,"hint":hint,"buildings":set(),"rows":[]}); group["rows"].append(r["_row"])
        if raw_build: group["buildings"].add(raw_build)
    reviews=[]; auto={}
    saved=rules.get("estates",{})
    for key,group in groups.items():
        raw=group["raw"]; hint=group["hint"]; rownums=group["rows"]; raw_buildings=sorted(group["buildings"])
        match_text=raw if not hint or phase_hint(raw) else raw+hint
        found=[]; method=""; conf=0
        saved_choice=saved.get(key,saved.get(raw,""))
        if saved_choice:
            saved_found=unique_estates([e for e in estates if estate_output_name(e)==saved_choice or e["楼盘"]==saved_choice])
            if len(saved_found)==1: found=saved_found; method="已确认规则"; conf=1
        exact_found=unique_estates(exact.get(estate_norm(match_text),[]))
        compact_found=unique_estates(compact.get(estate_norm(match_text,True),[]))
        if not found and len(exact_found)==1: found=exact_found; method="楼盘加期名精确匹配" if not hint else "结合栋座期名匹配"; conf=1
        if not found and len(compact_found)==1: found=compact_found; method="去除无意义差异后匹配" if not hint else "结合栋座期名匹配"; conf=.99
        if found:
            prefix=community_building_prefix(raw,found[0],unit_index)
            auto[key]={"estate":found[0],"method":method,"confidence":conf,"community_building_prefix":prefix}
            continue
        candidates=[]
        for e in estates:
            prefix=community_building_prefix(raw,e,unit_index)
            names=set(estate_match_names(e))
            for p in building_courtyard_prefixes(unit_index.get(e["楼盘"],{})):
                names.update(name+p for name in estate_match_names(e))
            score=max(SequenceMatcher(None,estate_norm(match_text,True),estate_norm(name,True)).ratio() for name in names)
            if score>=.35: candidates.append({**e,"score":round(score,3),"community_building_prefix":prefix})
        candidate_map={}
        for candidate in candidates:
            candidate_key=(candidate.get("match_name",estate_output_name(candidate)),candidate.get("行政区",""),candidate.get("商圈",""))
            if candidate_key not in candidate_map or candidate["score"]>candidate_map[candidate_key]["score"]: candidate_map[candidate_key]=candidate
        candidates=sorted(candidate_map.values(),key=lambda x:x["score"],reverse=True)[:5]
        reviews.append({"key":key,"raw":raw,"phaseHint":hint,"buildings":raw_buildings,"rows":rownums,"candidates":candidates,"selected":""})
    task["estate_auto"]=auto; task["estate_reviews"]=reviews
    return reviews, []

def stable_grade(seed): return ["A类","B类","C类"][int(hashlib.sha256(seed.encode("utf-8")).hexdigest(),16)%3]

def numeric_core(v):
    s=unicodedata.normalize("NFKC",str("" if v is None else v)); cn={"零":"0","〇":"0","一":"1","二":"2","三":"3","四":"4","五":"5","六":"6","七":"7","八":"8","九":"9"}
    for a,b in cn.items(): s=s.replace(a,b)
    m=re.search(r"\d+",s)
    return str(int(m.group())) if m else ""

def building_core(v):
    s=unicodedata.normalize("NFKC",str("" if v is None else v)); cn={"零":"0","〇":"0","一":"1","二":"2","三":"3","四":"4","五":"5","六":"6","七":"7","八":"8","九":"9"}
    for a,b in cn.items(): s=s.replace(a,b)
    building_numbers=re.findall(r"(\d+)\s*号?\s*(?:楼|栋|座|幢)",s)
    if building_numbers: return str(int(building_numbers[-1]))
    numbers=re.findall(r"\d+",s)
    return str(int(numbers[-1])) if numbers else ""

def load_unit_index():
    wb=openpyxl.load_workbook(RESOURCES/FILES["unit"],read_only=True,data_only=True); ws=wb[wb.sheetnames[0]]; rows=ws.iter_rows(values_only=True)
    heads=[str(x or "").strip() for x in next(rows)]; ix={h:i for i,h in enumerate(heads)}; index={}
    for r in rows:
        estate=str(r[ix["楼盘"]] or "").strip(); building=str(r[ix["座栋"]] or "").strip(); unit=str(r[ix["单元"]] or "").strip()
        if estate and building: index.setdefault(estate,{}).setdefault(building,set()).add(unit)
    return index

def confirmed_estates(task, selected=None):
    selected=selected or {}
    decisions=dict(task.get("estate_auto",{}))
    for rv in task.get("estate_reviews",[]):
        choice=selected.get(rv.get("key",rv["raw"]),selected.get(rv["raw"],rv.get("selected","")))
        if choice:
            e=next((c for c in rv["candidates"] if c.get("match_name",estate_output_name(c))==choice or c["楼盘"]==choice),None)
            if e: decisions[rv.get("key",rv["raw"])]= {"estate":e,"method":"人工确认","confidence":1,"community_building_prefix":e.get("community_building_prefix","")}
    return decisions

def building_review(task, estate_selected):
    for rv in task.get("estate_reviews",[]): rv["selected"]=estate_selected.get(rv.get("key",rv["raw"]),estate_selected.get(rv["raw"],""))
    decisions=confirmed_estates(task,estate_selected); index=load_unit_index(); src_by_target={}
    for src,tgts in task["mapping"].items():
        for t in (tgts if isinstance(tgts,list) else [tgts]):
            if t and t!="__ignore__": src_by_target.setdefault(t,[]).append(src)
    community_sources=src_by_target.get("小区",[]); building_sources=src_by_target.get("栋幢",[]); groups={}
    for row in task["rows"]:
        raw_build=" ".join(str(row.get(s) or "").strip() for s in building_sources).strip()
        raw_comm=" ".join(str(row.get(s) or "").strip() for s in community_sources).strip(); decision=decisions.get(community_decision_key(raw_comm,raw_build),decisions.get(raw_comm))
        if decision and raw_build:
            estate_info=decision["estate"]; estate=estate_info["楼盘"]; display=estate_output_name(estate_info); raw_build=merge_building_prefix(decision.get("community_building_prefix"),raw_build); groups.setdefault((estate,display,raw_build),[]).append(row["_row"])
        elif decision and decision.get("community_building_prefix"):
            estate_info=decision["estate"]; estate=estate_info["楼盘"]; display=estate_output_name(estate_info); raw_build=decision["community_building_prefix"]; groups.setdefault((estate,display,raw_build),[]).append(row["_row"])
    reviews=[]
    for (estate,display,raw),rownums in groups.items():
        buildings=list(index.get(estate,{})); exact=[b for b in buildings if norm(b)==norm(raw)]; core=[b for b in buildings if building_core(b) and building_core(b)==building_core(raw)]
        candidates=exact or core
        if len(candidates)==1: continue
        if not candidates:
            candidates=[b for _,b in sorted(((SequenceMatcher(None,norm(raw,True),norm(b,True)).ratio(),b) for b in buildings),reverse=True)[:8]]
        reviews.append({"key":display+"|"+raw,"estate":display,"base_estate":estate,"raw":raw,"rows":rownums,"candidates":candidates,"reason":"多个候选" if len(exact or core)>1 else "无法自动确认","selected":""})
    task["building_reviews"]=reviews
    return reviews

def save_rules(task):
    rules=jload(RULES_FILE,{"mappings":{},"estates":{},"buildings":{},"units":{}})
    rules.setdefault("mappings",{})[task["source_key"]]=task["mapping"]
    for x in task.get("estate_reviews",[]):
        if x.get("selected"): rules.setdefault("estates",{})[x.get("key",x["raw"])]=x["selected"]
    for x in task.get("building_reviews",[]):
        if x.get("selected"): rules.setdefault("buildings",{})[x["key"]]=x["selected"]
    jsave(RULES_FILE,rules)

def rule_snapshot(task):
    """生成随历史报告保存的规则快照，保证报告可复核且不依赖当前 rules.json。"""
    estate_rules=[]
    for key, decision in task.get("estate_auto", {}).items():
        estate=decision.get("estate", {})
        estate_rules.append({
            "原始键": key,
            "标准小区": estate_output_name(estate),
            "匹配方式": decision.get("method", ""),
            "置信度": decision.get("confidence", 0),
            "是否人工确认": decision.get("method") == "人工确认",
        })
    for review in task.get("estate_reviews", []):
        selected=review.get("selected", "")
        if selected:
            estate_rules.append({
                "原始键": review.get("key", review.get("raw", "")),
                "标准小区": selected,
                "匹配方式": "人工确认",
                "置信度": 1,
                "是否人工确认": True,
            })
    building_rules=[
        {
            "原始键": review.get("key", ""),
            "标准栋座": review.get("selected", ""),
            "是否人工确认": True,
        }
        for review in task.get("building_reviews", [])
        if review.get("selected")
    ]
    return {
        "版本": RULE_VERSION,
        "来源结构": task.get("source_key", ""),
        "字段映射": task.get("mapping", {}),
        "小区匹配": estate_rules,
        "栋座匹配": building_rules,
    }

def build_output(task, clean_only=False):
    kind=task["kind"]; info=template_info(kind); original_sig=task["template_signature"]
    if info["signature"]!=original_sig: raise ValueError("模板在任务开始后发生变化，请新建任务并重新确认字段")
    wb=openpyxl.load_workbook(info["path"]); ws=wb["房源"]
    if ws.max_row>=2: ws.delete_rows(2,ws.max_row-1)
    template_headers=info["headers"]
    headers=template_headers+["未匹配原因"]
    reason_col=len(headers)
    ws.cell(1,reason_col).value="未匹配原因"
    if reason_col>1:
        source_header=ws.cell(1,reason_col-1); reason_header=ws.cell(1,reason_col)
        if source_header.has_style: reason_header._style=copy(source_header._style)
    ws.column_dimensions[openpyxl.utils.get_column_letter(reason_col)].width=36
    src_by_target={}
    for s,ts in task["mapping"].items():
        for t in (ts if isinstance(ts,list) else [ts]):
            if t and t!="__ignore__": src_by_target.setdefault(t,[]).append(s)
    decisions=confirmed_estates(task)
    building_manual={x["key"]:x.get("selected","") for x in task.get("building_reviews",[]) if x.get("selected")}
    audit=[]; exceptions=[]; output=[]; now=utc_now_iso(); unit_index=load_unit_index()
    community_sources=src_by_target.get("小区",[])
    for source in task["rows"]:
        record={h:"" for h in headers}
        for target,sources in src_by_target.items():
            if target not in record or SYSTEM_BLANK.search(target): continue
            values=[source.get(s) for s in sources if source.get(s) not in (None,"")]
            if values: record[target]=" ".join(str(v) for v in values)
        layout_source=next((source.get(name) for name in source if aux_header(name) in ("房型","户型","居室","房屋户型") and source.get(name) not in (None,"")),"")
        layout=parse_layout(layout_source)
        for field,value in layout.items():
            if field not in record: continue
            current=str(record.get(field) or "").strip()
            if current and not re.search(r"[房室厅厨卫阳台]",current): continue
            if current!=str(value):
                audit.append({"原始行号":source["_row"],"房源编号":record.get("房源编号",""),"字段名称":field,"原始值":current or layout_source,"最终值":value,"使用的规则":"房型拆分","字典来源":"原始房型字段","匹配依据":layout_source,"置信度":1,"是否人工确认":False,"操作时间":now,"规则版本":RULE_VERSION})
            record[field]=value
        raw_comm=" ".join(str(source.get(s) or "").strip() for s in community_sources).strip()
        source_build=" ".join(str(source.get(s) or "").strip() for s in src_by_target.get("栋幢",[])).strip()
        decision=decisions.get(community_decision_key(raw_comm,source_build),decisions.get(raw_comm)); blocking=[]
        if decision:
            e=decision["estate"]
            estate_values={"小区":estate_output_name(e),"城区":e["行政区"],"商圈":e["商圈"]}
            for field,final_value in estate_values.items():
                if field in record and record[field]!=final_value:
                    audit.append({"原始行号":source["_row"],"房源编号":record.get("房源编号",""),"字段名称":field,"原始值":record[field],"最终值":final_value,"使用的规则":decision["method"],"字典来源":FILES["estate"],"匹配依据":raw_comm,"置信度":decision["confidence"],"是否人工确认":decision["method"]=="人工确认","操作时间":now,"规则版本":RULE_VERSION})
                    record[field]=final_value
            buildings=unit_index.get(e["楼盘"],{}); original_build=str(record.get("栋幢") or "").strip(); raw_build=merge_building_prefix(decision.get("community_building_prefix"),original_build); chosen_build=""
            if raw_build!=original_build:
                audit.append({"原始行号":source["_row"],"房源编号":record.get("房源编号",""),"字段名称":"栋幢","原始值":original_build,"最终值":raw_build,"使用的规则":"从小区名拆分几号院","字典来源":FILES["unit"],"匹配依据":raw_comm,"置信度":1,"是否人工确认":False,"操作时间":now,"规则版本":RULE_VERSION})
                record["栋幢"]=raw_build
            manual_build=building_manual.get(building_decision_key(e,raw_build),building_manual.get(e["楼盘"]+"|"+raw_build,""))
            exact_build=[b for b in buildings if norm(b)==norm(raw_build)]
            core_build=[b for b in buildings if building_core(b) and building_core(b)==building_core(raw_build)]
            candidates=exact_build or core_build
            if manual_build in buildings: chosen_build=manual_build
            elif raw_build and len(candidates)==1: chosen_build=candidates[0]
            elif raw_build and len(candidates)>1: blocking.append("栋座存在多个冲突候选")
            elif raw_build: blocking.append("栋座无法确认")
            if chosen_build:
                if record.get("栋幢")!=chosen_build:
                    audit.append({"原始行号":source["_row"],"房源编号":record.get("房源编号",""),"字段名称":"栋幢","原始值":record.get("栋幢",""),"最终值":chosen_build,"使用的规则":"栋座人工确认" if manual_build else "楼盘内唯一栋座规范","字典来源":FILES["unit"],"匹配依据":building_core(raw_build),"置信度":1,"是否人工确认":bool(manual_build),"操作时间":now,"规则版本":RULE_VERSION})
                    record["栋幢"]=chosen_build
                raw_unit=str(record.get("单元") or "").strip(); units=buildings[chosen_build]
                unit_values=sorted({str(u).strip() for u in units if str("" if u is None else u).strip()})
                zero_only=len(unit_values)==1 and numeric_core(unit_values[0])=="0"
                exact_unit=[u for u in units if norm(u)==norm(raw_unit)]; core_unit=[u for u in units if numeric_core(u) and numeric_core(u)==numeric_core(raw_unit)]; uc=exact_unit or core_unit
                if zero_only or (raw_unit and len(uc)==1):
                    chosen_unit=unit_values[0] if zero_only else uc[0]
                    if record.get("单元")!=chosen_unit:
                        audit.append({"原始行号":source["_row"],"房源编号":record.get("房源编号",""),"字段名称":"单元","原始值":record.get("单元",""),"最终值":chosen_unit,"使用的规则":"栋座仅有零单元" if zero_only else "楼盘栋座内唯一单元规范","字典来源":FILES["unit"],"匹配依据":numeric_core(raw_unit),"置信度":1,"是否人工确认":False,"操作时间":now,"规则版本":RULE_VERSION})
                        record["单元"]=chosen_unit
                elif raw_unit and len(uc)>1: blocking.append("单元存在多个冲突候选")
                elif raw_unit: blocking.append("单元无法确认")
        else: blocking.append("楼盘无法确认")
        price="总价(万)" if kind=="sale" else "月租金"
        if record.get(price) in (None,""): blocking.append(f"完全缺失{price}")
        record["委托日期"]=task["entrust_date"]
        seed=str(record.get("房源编号") or "".join(str(record.get(x) or "") for x in ("小区","栋幢","单元","房号","业主联系方式")))
        if record.get("等级") in (None,""):
            assigned_grade=stable_grade(seed); record["等级"]=assigned_grade
            audit.append({"原始行号":source["_row"],"房源编号":record.get("房源编号",""),"字段名称":"等级","原始值":"","最终值":assigned_grade,"使用的规则":"缺失等级自动分配","字典来源":"系统规则","匹配依据":seed,"置信度":1,"是否人工确认":False,"操作时间":now,"规则版本":RULE_VERSION})
        for h in headers:
            if SYSTEM_BLANK.search(h): record[h]=""
        record["未匹配原因"]="；".join(dict.fromkeys(blocking))
        if blocking: exceptions.append({"原始行号":source["_row"],"房源编号":record.get("房源编号",""),"原始状态":next((source.get(s) for s in source if "状态" in s),""),"异常":record["未匹配原因"]})
        if not clean_only or not blocking: output.append(record)
    for ri,rec in enumerate(output,2):
        for ci,h in enumerate(headers,1):
            cell=ws.cell(ri,ci); val=rec[h]
            if h=="委托日期":
                try: cell.value=datetime.strptime(val,"%Y-%m-%d").date(); cell.number_format="yyyy-mm-dd"
                except: cell.value=val
            else: cell.value=val
            if TEXT_FIELD.search(h): cell.number_format="@"
            if h=="未匹配原因": cell.alignment=copy(cell.alignment); cell.alignment=cell.alignment.copy(wrap_text=True,vertical="top")
        if ri>2:
            for ci in range(1,len(headers)+1):
                src=ws.cell(2,ci); dst=ws.cell(ri,ci)
                if src.has_style:
                    dst._style=copy(src._style); dst.number_format=src.number_format if not TEXT_FIELD.search(headers[ci-1]) else "@"
    stamp=datetime.now().strftime("%Y%m%d_%H%M%S"); prefix="售房" if kind=="sale" else "租房"
    outdir=OUTPUTS/task["id"]; outdir.mkdir(parents=True,exist_ok=True)
    outpath=outdir/f"{prefix}清洗导入_{stamp}.xlsx"; wb.save(outpath)
    report={
        "原始文件名": task["original_name"],
        "类型": "出售" if kind=="sale" else "出租",
        "任务开始时间": task["started"],
        "委托日期": task["entrust_date"],
        "原始记录数": len(task["rows"]),
        "最终输出记录数": len(output),
        "阻断异常数量": len(exceptions),
        "审计记录数": len(audit),
        "规则版本": RULE_VERSION,
        "完成时间": utc_now_iso(),
        "清洗规则": rule_snapshot(task),
        "异常详情": exceptions,
        "审计记录": audit,
    }
    storage = get_storage()
    artifacts = {}
    output_url = None
    download_url = None
    output_object_key = None
    if storage is not None:
        if task.get("source_artifact"):
            artifacts["原始上传文件"] = task["source_artifact"]
        artifact_paths = [outpath]
        for artifact_path in artifact_paths:
            if artifact_path.exists():
                object_key = f"housing-cleaner/{task['id']}/{artifact_path.name}"
                artifacts[artifact_path.name] = storage.upload_file(artifact_path, object_key)
        report["OSS文件"] = artifacts
        output_info = artifacts[outpath.name]
        output_url = str(output_info["url"])
        output_object_key = str(output_info["objectKey"])
        shutil.rmtree(outdir)
        source_path = task.get("source_path")
        if source_path and Path(source_path).exists():
            Path(source_path).unlink()
        if task.get("path") and Path(task["path"]) != Path(source_path or "") and Path(task["path"]).exists():
            Path(task["path"]).unlink()
        download_id = None
        output_file = None
    else:
        (outdir / "清洗报告.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), "utf-8")
        download_id = register_download(outpath)
        try:
            output_file = str(outpath.relative_to(ROOT))
        except ValueError:
            output_file = str(outpath)

    # Uploads are isolated under data/uploads/<task_id>/; remove the
    # temporary task directory after export in both local and OSS modes.
    task_upload_dir = UPLOADS / task["id"]
    if task_upload_dir.is_dir():
        shutil.rmtree(task_upload_dir, ignore_errors=True)

    source_info = task.get("source_artifact") or {}
    report_id = save_report(
        task["id"],
        kind,
        report,
        output_file,
        len(audit),
        str(source_info.get("url")) if source_info.get("url") else None,
        str(source_info.get("objectKey")) if source_info.get("objectKey") else None,
        output_url,
        output_object_key,
    )
    if output_object_key:
        download_url = f"/api/reports/{report_id}/download"
    task["last_export"]={"file":output_file,"downloadId":download_id,"ossUrl":output_url,"downloadUrl":download_url,"ossObjectKey":output_object_key,"artifacts":artifacts,"report":report,"audit":len(audit),"exceptions":exceptions}
    return task["last_export"]

class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args): print("[%s] %s"%(self.log_date_time_string(),fmt%args))
    def send_json(self,obj,status=200):
        body=json.dumps(obj,ensure_ascii=False,default=str).encode(); self.send_response(status); self.send_header("Content-Type","application/json; charset=utf-8"); self.send_header("Content-Length",str(len(body))); self.end_headers(); self.wfile.write(body)
    def body(self):
        n=int(self.headers.get("Content-Length",0)); return json.loads(self.rfile.read(n) or b"{}")
    def do_GET(self):
        path=urlparse(self.path).path
        if path=="/api/status": return self.send_json({"files":file_status(),"tasks":len(TASKS)})
        if path.startswith("/download/"):
            token=path[len("/download/"):]
            fp=resolve_download(token)
            if fp is None:
                rel=unquote(token); fp=(ROOT/rel).resolve()
                if ROOT.resolve() not in fp.parents:
                    self.send_response(302); self.send_header("Location","/?download=invalid"); self.end_headers(); return
                if not fp.exists():
                    parent=fp.parent; stamp=re.search(r"(\d{8}_\d{6})\.xlsx$",fp.name,re.I)
                    matches=list(parent.glob(f"*{stamp.group(1)}.xlsx")) if stamp and parent.exists() else []
                    if len(matches)==1: fp=matches[0].resolve()
            if not fp.exists():
                self.send_response(302); self.send_header("Location","/?download=missing"); self.end_headers(); return
            from urllib.parse import quote
            data=fp.read_bytes(); self.send_response(200); self.send_header("Content-Type","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"); self.send_header("Content-Disposition",f"attachment; filename*=UTF-8''{quote(fp.name)}"); self.send_header("Content-Length",str(len(data))); self.end_headers(); return self.wfile.write(data)
        fp=STATIC/("index.html" if path=="/" else path.lstrip("/"))
        if fp.exists():
            data=fp.read_bytes(); mime="text/html" if fp.suffix==".html" else "text/css" if fp.suffix==".css" else "application/javascript"
            self.send_response(200); self.send_header("Content-Type",mime+"; charset=utf-8"); self.send_header("Cache-Control","no-store, max-age=0"); self.send_header("Content-Length",str(len(data))); self.end_headers(); return self.wfile.write(data)
        self.send_error(404)
    def do_POST(self):
        try:
            path=urlparse(self.path).path
            if path=="/api/task":
                b=self.body(); kind=b.get("kind"); missing=[x["name"] for x in file_status() if not x["exists"]]
                if missing: raise ValueError("缺少基础文件："+"、".join(missing))
                if kind not in ("sale","rent"): raise ValueError("请先选择出售或出租")
                info=template_info(kind); tid=uuid.uuid4().hex[:12]; TASKS[tid]={"id":tid,"kind":kind,"started":utc_now_iso(),"entrust_date":str(date.today()),"template_signature":info["signature"]}
                return self.send_json({"taskId":tid,"headers":info["headers"]})
            if path=="/api/upload":
                tid=self.headers.get("X-Task-Id"); name=Path(unquote(self.headers.get("X-Filename","source.xlsx"))).name; task=get_task(tid); ext=Path(name).suffix.lower()
                if ext not in (".csv",".xls",".xlsx"): raise ValueError("仅支持 CSV、XLS、XLSX")
                task_upload_dir=UPLOADS/tid; task_upload_dir.mkdir(parents=True, exist_ok=True); fp=task_upload_dir/name; fp.write_bytes(self.rfile.read(int(self.headers.get("Content-Length",0)))); fp,sheets=inspect_upload(fp)
                task.update({"path":fp,"original_name":name,"sheets":sheets}); return self.send_json({"sheets":sheets,"received":{"name":name,"size":fp.stat().st_size,"sizeMB":round(fp.stat().st_size/1024/1024,1)}})
            if path=="/api/select-sheet":
                b=self.body(); task=get_task(b.get("taskId")); headers,rows=read_table(task["path"],b["sheet"],int(b["headerRow"])); task.update({"sheet":b["sheet"],"header_row":int(b["headerRow"]),"headers":headers,"rows":rows,"source_key":hashlib.sha256("|".join(map(norm,headers)).encode()).hexdigest()})
                target=template_info(task["kind"])["headers"]; return self.send_json({"headers":headers,"rows":len(rows),"suggestions":mapping_suggestions(headers,target,rows,task["source_key"]),"targets":target})
            if path=="/api/mapping":
                b=self.body(); task=get_task(b.get("taskId")); task["mapping"]=b["mapping"]; reviews,errors=estate_review(task); return self.send_json({"reviews":reviews,"errors":errors,"autoCount":len(task.get("estate_auto",{}))})
            if path=="/api/review":
                b=self.body(); task=get_task(b.get("taskId")); selected=b.get("selected",{})
                for rv in task.get("estate_reviews",[]): rv["selected"]=selected.get(rv.get("key",rv["raw"]),selected.get(rv["raw"],""))
                task["entrust_date"]=b.get("entrustDate",task["entrust_date"]); save_rules(task); return self.send_json({"ok":True})
            if path=="/api/building-review":
                b=self.body(); task=get_task(b.get("taskId")); reviews=building_review(task,b.get("selected",{})); return self.send_json({"reviews":reviews})
            if path=="/api/building-confirm":
                b=self.body(); task=get_task(b.get("taskId")); selected=b.get("selected",{})
                for rv in task.get("building_reviews",[]): rv["selected"]=selected.get(rv["key"],"")
                task["entrust_date"]=b.get("entrustDate",task["entrust_date"])
                if b.get("persistRules",True): save_rules(task)
                return self.send_json({"ok":True})
            if path=="/api/export":
                b=self.body(); return self.send_json(build_output(get_task(b.get("taskId")),bool(b.get("cleanOnly"))))
            self.send_error(404)
        except Exception as e: self.send_json({"error":str(e)},400)
