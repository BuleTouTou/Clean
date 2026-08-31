from __future__ import annotations

import io
import os
from datetime import date

import httpx
import openpyxl
import pytest


BASE_URL = os.getenv("CONTAINER_BASE_URL")


@pytest.mark.skipif(not BASE_URL, reason="设置 CONTAINER_BASE_URL 后执行容器集成测试")
def test_postgresql_minio_container_flow() -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "原始数据"
    sheet.append(["北京房源导出表"])
    sheet.append(["楼盘", "行政区", "板块", "栋号", "单元", "室号", "售价", "业主电话"])
    sheet.append(["连心园", "朝阳", "常营", "1号楼", "1单元", "101", 500, "0013800000000"])
    payload = io.BytesIO()
    workbook.save(payload)

    with httpx.Client(base_url=BASE_URL, timeout=60) as client:
        response = client.post("/api/task", json={"kind": "sale"})
        response.raise_for_status()
        task_id = response.json()["taskId"]

        response = client.post(
            "/api/upload",
            content=payload.getvalue(),
            headers={
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "X-Task-Id": task_id,
                "X-Filename": "container-smoke.xlsx",
            },
        )
        response.raise_for_status()

        response = client.post(
            "/api/select-sheet",
            json={"taskId": task_id, "sheet": "原始数据", "headerRow": 2},
        )
        response.raise_for_status()
        selected = response.json()
        mapping = {item["source"]: (item["target"] or "__ignore__") for item in selected["suggestions"]}
        mapping.update(
            {
                "楼盘": "小区",
                "行政区": "城区",
                "板块": "商圈",
                "栋号": "栋幢",
                "单元": "单元",
                "室号": "房号",
                "售价": "总价(万)",
                "业主电话": "业主联系方式",
            }
        )

        response = client.post("/api/mapping", json={"taskId": task_id, "mapping": mapping})
        response.raise_for_status()
        response = client.post(
            "/api/review",
            json={"taskId": task_id, "selected": {}, "entrustDate": str(date.today())},
        )
        response.raise_for_status()
        response = client.post("/api/export", json={"taskId": task_id, "cleanOnly": False})
        response.raise_for_status()
        exported = response.json()

        assert exported["ossObjectKey"]
        assert exported["downloadUrl"].startswith("/api/reports/")
        download = client.get(exported["downloadUrl"])
        download.raise_for_status()
        result = openpyxl.load_workbook(io.BytesIO(download.content))
        assert result["房源"].max_row == 2
