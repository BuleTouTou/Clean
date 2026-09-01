from __future__ import annotations

import io
from datetime import date

import openpyxl
import pytest
from fastapi.testclient import TestClient

from backend import core
from backend.app import app


def estate(phase: str) -> dict[str, str]:
    return {"楼盘": "幸福里", "别名": "", "行政区": "朝阳", "商圈": "常营", "期名": phase}


def unit_catalog(records: list[tuple[str, str, str, str]]) -> dict:
    by_estate: dict = {}
    by_estate_norm: dict = {}
    by_estate_phase: dict = {}
    phases_by_estate: dict = {}
    phase_labels: dict = {}
    for estate_name, phase, building, unit in records:
        estate_key = core.estate_norm(estate_name, True)
        phase_key = core.estate_norm(phase, True)
        by_estate.setdefault(estate_name, {}).setdefault(building, set()).add(unit)
        by_estate_norm.setdefault(estate_key, {}).setdefault(building, set()).add(unit)
        by_estate_phase.setdefault((estate_key, phase_key), {}).setdefault(building, set()).add(unit)
        phases_by_estate.setdefault(estate_key, {}).setdefault(phase_key, {}).setdefault(building, set()).add(unit)
        if phase_key:
            phase_labels.setdefault(estate_key, {})[phase_key] = phase
    return {
        "by_estate": by_estate,
        "by_estate_norm": by_estate_norm,
        "by_estate_phase": by_estate_phase,
        "phases_by_estate": phases_by_estate,
        "phase_labels": phase_labels,
    }


def task(*buildings: str) -> dict:
    return {
        "mapping": {"楼盘": "小区", "栋号": "栋幢"},
        "rows": [
            {"_row": index + 2, "楼盘": "幸福里", "栋号": building}
            for index, building in enumerate(buildings)
        ],
    }


def install_dictionary(monkeypatch: pytest.MonkeyPatch, catalog: dict) -> None:
    monkeypatch.setattr(core, "load_estates", lambda: [estate("一期"), estate("二期")])
    monkeypatch.setattr(core, "load_unit_catalog", lambda: catalog)
    monkeypatch.setattr(core, "jload", lambda _path, default: default)


def test_buildings_split_same_community_into_unique_phases(monkeypatch: pytest.MonkeyPatch) -> None:
    catalog = unit_catalog(
        [
            ("幸福里", "一期", "1号楼", "1单元"),
            ("幸福里", "一期", "15号楼", "1单元"),
            ("幸福里", "二期", "16号楼", "1单元"),
            ("幸福里", "二期", "30号楼", "1单元"),
        ]
    )
    install_dictionary(monkeypatch, catalog)

    current_task = task("1号楼", "16号楼")
    reviews, errors = core.estate_review(current_task)

    assert reviews == []
    assert errors == []
    decisions = core.confirmed_estates(current_task)
    assert core.estate_output_name(decisions[core.community_decision_key("幸福里", "1号楼")]["estate"]) == "幸福里一期"
    assert core.estate_output_name(decisions[core.community_decision_key("幸福里", "16号楼")]["estate"]) == "幸福里二期"
    assert {decision["method"] for decision in decisions.values()} == {"栋座唯一对应期名"}


def test_ambiguous_buildings_are_grouped_and_batch_is_guarded(monkeypatch: pytest.MonkeyPatch) -> None:
    catalog = unit_catalog(
        [
            ("幸福里", "一期", "1号楼", "1单元"),
            ("幸福里", "一期", "2号楼", "1单元"),
            ("幸福里", "二期", "1号楼", "1单元"),
            ("幸福里", "二期", "2号楼", "1单元"),
        ]
    )
    install_dictionary(monkeypatch, catalog)

    current_task = task("1号楼", "2号楼")
    reviews, _errors = core.estate_review(current_task)

    assert len(reviews) == 1
    group = reviews[0]
    assert group["buildingCount"] == 2
    assert {item["rawBuilding"] for item in group["buildings"]} == {"1号楼", "2号楼"}
    assert all(candidate["coverageCount"] == 2 for candidate in group["candidates"])
    assert all(candidate["canApplyAll"] is False for candidate in group["candidates"])

    candidate_id = group["candidates"][0]["id"]
    selected = {item["key"]: candidate_id for item in group["buildings"]}
    with pytest.raises(ValueError, match="安全条件"):
        core.apply_estate_selections(
            current_task,
            selected,
            {key: "batch-all" for key in selected},
        )

    core.apply_estate_selections(
        current_task,
        selected,
        {key: "batch-compatible" for key in selected},
    )
    assert {
        decision["method"] for decision in core.confirmed_estates(current_task).values()
    } == {"批量确认兼容栋座"}

    with pytest.raises(ValueError, match="候选无效"):
        core.apply_estate_selections(current_task, {group["buildings"][0]["key"]: "tampered"})


def test_safe_candidate_can_apply_to_all_buildings() -> None:
    candidate = {
        **core.prepare_estate(estate("一期")),
        "buildingState": "unique",
        "matchedBuildings": ["1号楼"],
        "buildingPhases": ["一期"],
        "score": 0.8,
        "community_building_prefix": "",
    }
    reviews = []
    for row_number, building in ((2, "1号楼"), (3, "2号楼")):
        reviews.append(
            {
                "key": core.community_decision_key("幸福家园", building),
                "raw": "幸福家园",
                "rawBuilding": building,
                "rows": [row_number],
                "candidates": [{**candidate, "matchedBuildings": [building]}],
                "selected": "",
                "selectionMethod": "",
            }
        )
    groups = core.group_estate_reviews(reviews)
    current_task = {"estate_reviews": reviews, "estate_review_groups": groups, "estate_auto": {}}
    group = groups[0]
    summary = group["candidates"][0]

    assert summary["canApplyAll"] is True
    selected = {item["key"]: summary["id"] for item in group["buildings"]}
    methods = {key: "batch-all" for key in selected}
    core.apply_estate_selections(current_task, selected, methods)
    assert {decision["method"] for decision in core.confirmed_estates(current_task).values()} == {"批量确认全部"}


def test_export_uses_different_phases_for_same_community(monkeypatch: pytest.MonkeyPatch) -> None:
    catalog = unit_catalog(
        [
            ("幸福里", "一期", "1号楼", "1单元"),
            ("幸福里", "二期", "16号楼", "1单元"),
        ]
    )
    install_dictionary(monkeypatch, catalog)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "原始数据"
    sheet.append(["楼盘", "栋号", "单元", "室号", "售价"])
    sheet.append(["幸福里", "1号楼", "1单元", "101", 500])
    sheet.append(["幸福里", "16号楼", "1单元", "1601", 600])
    payload = io.BytesIO()
    workbook.save(payload)

    with TestClient(app) as client:
        login = client.post("/api/auth/login", json={"username": "admin", "password": "test-admin-password"})
        assert login.status_code == 200
        created = client.post("/api/task", json={"kind": "sale"})
        assert created.status_code == 200
        task_id = created.json()["taskId"]
        upload = client.post(
            "/api/upload",
            content=payload.getvalue(),
            headers={
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "X-Task-Id": task_id,
                "X-Filename": "phase-split.xlsx",
            },
        )
        assert upload.status_code == 200
        selected_sheet = client.post(
            "/api/select-sheet",
            json={"taskId": task_id, "sheet": "原始数据", "headerRow": 1},
        )
        assert selected_sheet.status_code == 200
        mapping = client.post(
            "/api/mapping",
            json={
                "taskId": task_id,
                "mapping": {
                    "楼盘": "小区",
                    "栋号": "栋幢",
                    "单元": "单元",
                    "室号": "房号",
                    "售价": "总价(万)",
                },
            },
        )
        assert mapping.status_code == 200
        assert mapping.json()["reviews"] == []
        assert mapping.json()["autoCount"] == 2
        building_review = client.post(
            "/api/building-review",
            json={"taskId": task_id, "selected": {}, "selectionMethods": {}},
        )
        assert building_review.status_code == 200
        building_confirm = client.post(
            "/api/building-confirm",
            json={
                "taskId": task_id,
                "selected": {},
                "entrustDate": str(date.today()),
                "persistRules": False,
            },
        )
        assert building_confirm.status_code == 200
        exported = client.post("/api/export", json={"taskId": task_id, "cleanOnly": False})
        assert exported.status_code == 200
        downloaded = client.get(exported.json()["downloadUrl"])
        assert downloaded.status_code == 200

    result = openpyxl.load_workbook(io.BytesIO(downloaded.content), data_only=True)
    output_sheet = result["房源"]
    headers = [str(cell.value or "") for cell in output_sheet[1]]
    community_column = headers.index("小区") + 1
    assert output_sheet.cell(2, community_column).value == "幸福里一期"
    assert output_sheet.cell(3, community_column).value == "幸福里二期"
