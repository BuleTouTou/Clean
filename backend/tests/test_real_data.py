import json
import os
import re
import sys
from datetime import date
from pathlib import Path
from urllib.parse import quote

import openpyxl
from fastapi.testclient import TestClient

from backend import core as server
from backend.app import app


def req(client: TestClient, url: str, obj=None, headers=None, raw=None):
    headers = headers or {}
    if obj is not None:
        headers["Content-Type"] = "application/json"
        response = client.post(url, content=json.dumps(obj, ensure_ascii=False).encode("utf-8"), headers=headers)
    elif raw is not None:
        response = client.post(url, content=raw, headers=headers)
    else:
        response = client.get(url, headers=headers)
    assert response.status_code < 400, response.text
    return response.json()


def main(path: str):
    source = Path(path)
    assert source.exists(), f"输入文件不存在：{source}"
    with TestClient(app) as client:
        req(client, "/api/auth/login", {"username": "admin", "password": os.environ["ADMIN_INITIAL_PASSWORD"]})
        task = req(client, "/api/task", {"kind": "sale"})
        task_id = task["taskId"]
        upload = req(client, "/api/upload", headers={"X-Task-Id": task_id, "X-Filename": quote(source.name)}, raw=source.read_bytes())
        sheet = upload["sheets"][0]
        reported_rows = sheet["rows"]
        selected = req(client, "/api/select-sheet", {"taskId": task_id, "sheet": sheet["name"], "headerRow": sheet["headerRow"]})
        targets = set(selected["targets"])
        mapping = {item["source"]: (item["source"] if item["source"] in targets else (item["target"] or "__ignore__")) for item in selected["suggestions"]}
        estate = req(client, "/api/mapping", {"taskId": task_id, "mapping": mapping})
        estate_choices = {}
        for group in estate["reviews"]:
            candidate_ids = {candidate["id"] for candidate in group["candidates"]}
            for building_item in group["buildings"]:
                choice = next((candidate_id for candidate_id in building_item["candidateIds"] if candidate_id in candidate_ids), "")
                if choice:
                    estate_choices[building_item["key"]] = choice
        building = req(client, "/api/building-review", {"taskId": task_id, "selected": estate_choices})
        building_choices = {item["key"]: item["candidates"][0] for item in building["reviews"] if item["candidates"]}
        req(client, "/api/building-confirm", {"taskId": task_id, "selected": building_choices, "entrustDate": str(date.today()), "persistRules": False})
        exported = req(client, "/api/export", {"taskId": task_id, "cleanOnly": False})

    output_path = server.ROOT / exported["file"]
    workbook = openpyxl.load_workbook(output_path, read_only=True, data_only=False)
    sheet = workbook["房源"]
    headers = [str(cell.value or "") for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    community_col = headers.index("小区")
    reason_col = headers.index("未匹配原因")
    output_rows = reason_rows = bracket_phase_rows = courtyard_communities = leading_apostrophes = composite_layout_values = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in row):
            continue
        output_rows += 1
        community = str(row[community_col] or "")
        if row[reason_col] not in (None, ""):
            reason_rows += 1
        if re.search(r"[（(]\s*(?:第?[一二三四五六七八九十\d]+期|[一二三四五六七八九十\d]+区)\s*[）)]", community):
            bracket_phase_rows += 1
        if re.search(r"\d+号院$", community):
            courtyard_communities += 1
        leading_apostrophes += sum(isinstance(value, str) and value.startswith("'") for value in row)
        composite_layout_values += sum(isinstance(row[index], str) and re.search(r"[房室厅厨卫阳台]", row[index]) is not None for index in [headers.index(item) for item in ("室", "厅", "厨", "卫", "阳台")])
    result = {
        "input_file": source.name,
        "reported_rows": reported_rows,
        "output_rows": output_rows,
        "estate_auto": estate["autoCount"],
        "estate_reviews": len(estate["reviews"]),
        "building_reviews": len(building["reviews"]),
        "rows_with_reason": reason_rows,
        "phase_parentheses_remaining": bracket_phase_rows,
        "communities_still_ending_courtyard": courtyard_communities,
        "leading_apostrophes_remaining": leading_apostrophes,
        "composite_layout_values_remaining": composite_layout_values,
        "output": str(output_path),
        "report": exported["report"],
    }
    assert output_rows == exported["report"]["最终输出记录数"]
    assert bracket_phase_rows == 0
    assert leading_apostrophes == 0
    assert composite_layout_values == 0
    print(json.dumps(result, ensure_ascii=False))


def test_real_data_file():
    """通过 REAL_DATA_FILE 环境变量启用真实文件验收。"""
    configured = os.environ.get("REAL_DATA_FILE")
    if not configured:
        import pytest

        pytest.skip("设置 REAL_DATA_FILE 后执行真实文件验收")
    main(configured)


if __name__ == "__main__":
    if len(sys.argv) != 2:
        raise SystemExit("用法：uv run --project backend python backend/tests/test_real_data.py /path/to/source.xlsx")
    main(sys.argv[1])
