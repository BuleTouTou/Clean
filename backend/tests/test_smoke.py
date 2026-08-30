import json
from datetime import date

import openpyxl
from fastapi.testclient import TestClient

from backend import core as server
from backend.app import app


def req(client: TestClient, url: str, obj=None, headers=None, raw=None):
    headers = headers or {}
    if obj is not None:
        headers["Content-Type"] = "application/json"
        response = client.post(url, content=json.dumps(obj).encode(), headers=headers)
    elif raw is not None:
        response = client.post(url, content=raw, headers=headers)
    else:
        response = client.get(url, headers=headers)
    assert response.status_code < 400, response.text
    return response.json()


def main():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "原始数据"
    sheet.append(["北京房源导出表"])
    sheet.merge_cells("A1:H1")
    sheet.append(["楼盘", "行政区", "板块", "栋号", "单元", "室号", "售价", "业主电话", "录入人工号", "维护人工号", "房源等级"])
    sheet.append(["连心园", "朝阳", "常营", "1号楼", "1单元", "101", 500, "0013800000000", "0260616013", "260616013", "S级"])
    sheet.append(["绝对不存在小区", "海淀", "中关村", "2栋", "2", "202", "", "13900000000", "0260616010", "260616010", ""])

    source = server.DATA / "smoke.xlsx"
    source.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(source)
    try:
        with TestClient(app) as client:
            task = req(client, "/api/task", {"kind": "sale"})
            task_id = task["taskId"]
            upload = req(client, "/api/upload", headers={"X-Task-Id": task_id, "X-Filename": "smoke.xlsx"}, raw=source.read_bytes())
            selected = req(client, "/api/select-sheet", {"taskId": task_id, "sheet": "原始数据", "headerRow": 2})
            mapping = {item["source"]: (item["target"] or "__ignore__") for item in selected["suggestions"]}
            mapping.update({"楼盘": "小区", "行政区": "城区", "板块": "商圈", "栋号": "栋幢", "室号": "房号", "售价": "总价(万)", "业主电话": "业主联系方式", "单元": "单元", "录入人工号": "录入人工号", "维护人工号": "维护人工号"})
            review = req(client, "/api/mapping", {"taskId": task_id, "mapping": mapping})
            req(client, "/api/review", {"taskId": task_id, "selected": {}, "entrustDate": str(date.today())})
            exported = req(client, "/api/export", {"taskId": task_id, "cleanOnly": False})

        output_path = server.ROOT / exported["file"]
        result = openpyxl.load_workbook(output_path)
        output_sheet = result["房源"]
        assert [cell.value for cell in output_sheet[1]][:-1] == task["headers"]
        assert output_sheet.cell(1, output_sheet.max_column).value == "未匹配原因"
        expected_community = server.estate_output_name(next(e for e in server.load_estates() if e["楼盘"] == "连心园"))
        assert output_sheet.max_row == 3 and output_sheet["F2"].value == expected_community and output_sheet["B2"].is_date
        assert output_sheet["Y2"].value == "S级" and output_sheet["Y3"].value in ("A类", "B类", "C类") and output_sheet["AB2"].number_format == "@"
        assert str(output_sheet["AL2"].value) == "0260616013" and output_sheet["AL2"].number_format == "@"
        assert str(output_sheet["AP2"].value) == "260616013" and output_sheet["AP2"].number_format == "@"
        assert exported["report"]["阻断异常数量"] >= 1
        assert exported["report"]["清洗规则"]["字段映射"] == mapping
        assert any(output_sheet.cell(row, output_sheet.max_column).value for row in range(2, output_sheet.max_row + 1))
        print(json.dumps({"ok": True, "output": str(output_path), "review": len(review["reviews"]), "report": exported["report"]}, ensure_ascii=False))
    finally:
        source.unlink(missing_ok=True)


def test_smoke():
    main()


if __name__ == "__main__":
    main()
