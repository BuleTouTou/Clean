from __future__ import annotations

import csv, hashlib, io, json, os, re, shutil, subprocess, threading, time, unicodedata, uuid, webbrowser
from copy import copy
from datetime import date, datetime
from difflib import SequenceMatcher
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse, unquote

import openpyxl

ROOT = Path(__file__).resolve().parent
DATA = ROOT / "data"
UPLOADS = DATA / "uploads"
OUTPUTS = ROOT / "outputs"
RULES_FILE = DATA / "rules.json"
DOWNLOADS_FILE = DATA / "downloads.json"
STATIC = ROOT / "static"
for p in (DATA, UPLOADS, OUTPUTS): p.mkdir(exist_ok=True)

FILES = {
    "estate": "北京楼盘字典.xlsx",
    "unit": "北京-单元楼盘字典.xlsx",
    "sale": "售房数据导入模版.xlsx",
    "rent": "租房数据导入模版 .xlsx",
}
RULE_VERSION = "1.0.0"
TASKS: dict[str, dict] = {}
DOWNLOADS: dict[str, Path] = {}
LOCK = threading.RLock()

SYNONYMS = {
    "城区": ["行政区", "城区", "区域", "所属城区"], "商圈": ["片区", "商圈", "板块"],
    "小区": ["小区", "楼盘", "项目名称", "社区"], "栋幢": ["座栋", "栋幢", "楼栋", "栋号"],
    "房号": ["房号", "门牌号", "室号"], "面积": ["建筑面积", "建面", "面积", "面积㎡"],
    "总价(万)": ["售价", "出售价格", "总价"], "月租金": ["租价", "租金", "月租", "出租价格"],
    "业主姓名": ["业主", "房东", "产权人"], "业主联系方式": ["电话", "手机号", "联系方式", "业主电话"],
    "录入人": ["录入人姓名", "录入员工", "录入人员"], "维护人": ["维护人姓名", "维护员工", "维护人员"],
}
# 工号和部门代码不再无条件清空：只有用户明确映射且源值非空时才写入。
# 未映射字段仍保持模板空值，系统不会推测或生成任何代码。
SYSTEM_BLANK = re.compile(r"(?!)")
TEXT_FIELD = re.compile(r"(编号|电话|联系方式|工号|代码)$")

def jload(path: Path, default):
    try: return json.loads(path.read_text("utf-8"))
    except Exception: return default

def jsave(path: Path, obj): path.write_text(json.dumps(obj, ensure_ascii=False, indent=2), "utf-8")

def register_download(path: Path):
    token=uuid.uuid4().hex
    registry=jload(DOWNLOADS_FILE,{})
    registry[token]=str(path.resolve())
    jsave(DOWNLOADS_FILE,registry)
    DOWNLOADS[token]=path.resolve()
    return token

def resolve_download(token):
    fp=DOWNLOADS.get(token)
    if fp: return fp
    raw=jload(DOWNLOADS_FILE,{}).get(token)
    if raw:
        fp=Path(raw).resolve()
        if ROOT.resolve() in fp.parents:
            DOWNLOADS[token]=fp
            return fp
    return None

def get_task(task_id):
    task=TASKS.get(task_id)
    if task is None:
        raise ValueError("当前批次已因工具重启而失效，请返回首页重新创建批次")
    return task

def norm(v, compact=False):
    s = unicodedata.normalize("NFKC", str(v or "")).strip().replace("\r", "").replace("\n", "")
    s = s.casefold()
    if compact:
        s = re.sub(r"[（(][^）)]*[）)]", "", s)
        s = re.sub(r"[\s·•,，.。\-_—/\\#号座栋幢楼]", "", s)
    return s

def aux_header(v): return re.sub(r"[（(][^）)]*[）)]", "", norm(v)).strip()

def file_status():
    return [{"key": k, "name": n, "exists": (ROOT/n).exists(), "size": (ROOT/n).stat().st_size if (ROOT/n).exists() else 0,
             "mtime": (ROOT/n).stat().st_mtime if (ROOT/n).exists() else None} for k,n in FILES.items()]

def template_info(kind):
    p = ROOT / FILES[kind]
    wb = openpyxl.load_workbook(p, read_only=True, data_only=False)
    if "房源" not in wb.sheetnames: raise ValueError(f"模板 {p.name} 缺少‘房源’工作表")
    ws = wb["房源"]
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    if not all(headers): raise ValueError("模板表头包含空列")
    return {"path": p, "headers": headers, "signature": hashlib.sha256(("|".join(headers)+str(p.stat().st_mtime_ns)).encode()).hexdigest()}

def detect_header(rows):
    best = (0, -1)
    for i,row in enumerate(rows[:30]):
        vals = [str(x).strip() for x in row if x not in (None, "")]
        unique = len(set(vals)); textish = sum(any(ch.isalpha() or '\u4e00' <= ch <= '\u9fff' for ch in x) for x in vals)
        score = unique * 2 + textish - max(0, len(vals)-unique)*3
        if len(vals) >= 2 and score > best[1]: best = (i, score)
    return best[0]

def csv_rows(path):
    raw = path.read_bytes(); encoding = None
    for enc in ("utf-8-sig", "utf-8", "gb18030", "gbk"):
        try: text = raw.decode(enc); encoding = enc; break
        except UnicodeDecodeError: pass
    if encoding is None: raise ValueError("无法识别 CSV 编码")
    sample = text[:65536]
    try: dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|，")
    except csv.Error: dialect = csv.excel
    rows = list(csv.reader(io.StringIO(text), dialect))
    return rows, encoding, dialect.delimiter

def convert_xls(path):
    target = path.with_suffix(".xlsx")
    soffice = shutil.which("soffice")
    if soffice:
        subprocess.run([soffice,"--headless","--convert-to","xlsx","--outdir",str(path.parent),str(path)], check=True, timeout=120)
        return target
    raise ValueError("旧版 XLS 需要本机安装 LibreOffice；请另存为 XLSX 后重试")

def inspect_upload(path):
    ext = path.suffix.lower(); sheets=[]
    if ext == ".csv":
        rows, enc, delim = csv_rows(path); hi=detect_header(rows); width=max((len(r) for r in rows), default=0)
        sheets.append({"name":"CSV", "hidden":False, "rows":max(0,len(rows)-hi-1), "cols":width, "headerRow":hi+1,
                       "preview":[list(map(str,r[:20])) for r in rows[:5]], "encoding":enc, "delimiter":delim})
    else:
        if ext == ".xls": path=convert_xls(path)
        wb=openpyxl.load_workbook(path, read_only=False, data_only=True)
        for ws in wb.worksheets:
            sample=list(ws.iter_rows(min_row=1,max_row=min(ws.max_row,30),values_only=True)); hi=detect_header(sample)
            sheets.append({"name":ws.title,"hidden":ws.sheet_state!="visible","rows":max(0,ws.max_row-hi-1),"cols":ws.max_column,
                           "headerRow":hi+1,"preview":[["" if v is None else str(v) for v in r[:20]] for r in sample[:5]],
                           "merged":[str(x) for x in list(ws.merged_cells.ranges)[:50]]})
    return path, sheets

def read_table(path, sheet, header_row):
    if path.suffix.lower()==".csv":
        all_rows,_,_=csv_rows(path); row_iter=iter(all_rows)
    else:
        wb=openpyxl.load_workbook(path, read_only=True, data_only=True); ws=wb[sheet]; row_iter=ws.iter_rows(values_only=True)
    for _ in range(header_row-1): next(row_iter,None)
    header_values=next(row_iter,())
    raw=["" if x is None else str(x).strip() for x in header_values]
    while raw and raw[-1]=="": raw.pop()
    seen={}; headers=[]
    for i,h in enumerate(raw):
        h=h or f"未命名列{i+1}"; seen[h]=seen.get(h,0)+1; headers.append(h if seen[h]==1 else f"{h}_{seen[h]}")
    data=[]
    for rn,row in enumerate(row_iter, start=header_row+1):
        vals=list(row[:len(headers)])+[None]*max(0,len(headers)-len(row))
        if any(v not in (None,"") for v in vals): data.append({"_row":rn, **dict(zip(headers,vals))})
    return headers,data

def mapping_suggestions(headers, target, rows, source_key):
    rules=jload(RULES_FILE, {"mappings":{},"estates":{},"buildings":{},"units":{}})
    saved=rules.get("mappings",{}).get(source_key,{})
    syn={norm(x):k for k,vals in SYNONYMS.items() for x in vals}
    out=[]
    for h in headers:
        nh=norm(h); ah=aux_header(h); suggestion=""; confidence=0; reason="需人工选择"
        if h in target: suggestion=h; confidence=1; reason="字段名完全一致"
        elif nh in syn and syn[nh] in target: suggestion=syn[nh]; confidence=.98; reason="同义词完全匹配"
        elif h in saved and saved[h] in target: suggestion=saved[h]; confidence=.97; reason="已确认来源映射"
        elif ah in [norm(x) for x in target]: suggestion=target[[norm(x) for x in target].index(ah)]; confidence=.9; reason="去除单位后匹配"
        else:
            scored=sorted(((SequenceMatcher(None,ah,norm(x)).ratio(),x) for x in target),reverse=True)
            if scored and scored[0][0]>=.55: confidence,suggestion=scored[0]; reason="名称相似，仅供建议"
        samples=[]
        for r in rows:
            if r.get(h) not in (None,""):
                samples.append(str(r.get(h)))
                if len(samples)>=5: break
        out.append({"source":h,"target":suggestion,"confidence":round(confidence,2),"reason":reason,"samples":samples})
    return out

def load_estates():
    wb=openpyxl.load_workbook(ROOT/FILES["estate"],read_only=True,data_only=True); ws=wb[wb.sheetnames[0]]
    rows=ws.iter_rows(values_only=True); heads=[str(x or "").strip() for x in next(rows)]; ix={h:i for i,h in enumerate(heads)}
    result=[]
    for r in rows:
        if not r[ix["楼盘"]]: continue
        result.append({k:("" if r[ix[k]] is None else str(r[ix[k]]).strip()) for k in ("楼盘","别名","行政区","商圈","期名")})
    return result

def estate_review(task):
    mapping=task["mapping"]; rows=task["rows"]; estates=load_estates(); rules=jload(RULES_FILE,{})
    src_by_target={}
    for src,tgts in mapping.items():
        for t in (tgts if isinstance(tgts,list) else [tgts]):
            if t and t!="__ignore__": src_by_target.setdefault(t,[]).append(src)
    community_src=src_by_target.get("小区",[])
    if not community_src: return [], [{"type":"小区没有映射字段","count":len(rows)}]
    exact={}; alias={}; compact={}
    for e in estates:
        exact.setdefault(norm(e["楼盘"]),[]).append(e); compact.setdefault(norm(e["楼盘"],True),[]).append(e)
        for a in re.split(r"[,，、;/；|]",e["别名"]):
            if a.strip(): alias.setdefault(norm(a),[]).append(e); compact.setdefault(norm(a,True),[]).append(e)
    groups={}
    for r in rows:
        raw=" ".join(str(r.get(s) or "").strip() for s in community_src).strip(); groups.setdefault(raw,[]).append(r["_row"])
    reviews=[]; auto={}
    saved=rules.get("estates",{})
    for raw,rownums in groups.items():
        found=[]; method=""; conf=0
        if raw in saved: found=[e for e in estates if e["楼盘"]==saved[raw]]; method="已确认规则"; conf=1
        if not found and len(exact.get(norm(raw),[]))==1: found=exact[norm(raw)]; method="标准楼盘精确匹配"; conf=1
        if not found and len(alias.get(norm(raw),[]))==1: found=alias[norm(raw)]; method="别名精确匹配"; conf=1
        if not found and len(compact.get(norm(raw,True),[]))==1: found=compact[norm(raw,True)]; method="去除无意义差异后匹配"; conf=.99
        if found: auto[raw]={"estate":found[0],"method":method,"confidence":conf}; continue
        candidates=[]
        for e in estates:
            score=max(SequenceMatcher(None,norm(raw,True),norm(e["楼盘"],True)).ratio(),
                      max([SequenceMatcher(None,norm(raw,True),norm(a,True)).ratio() for a in re.split(r"[,，、;/；|]",e["别名"]) if a] or [0]))
            if score>=.35: candidates.append({**e,"score":round(score,3)})
        candidates=sorted(candidates,key=lambda x:x["score"],reverse=True)[:5]
        reviews.append({"raw":raw,"rows":rownums,"candidates":candidates,"selected":""})
    task["estate_auto"]=auto; task["estate_reviews"]=reviews
    return reviews, []

def stable_grade(seed): return ["A类","B类","C类"][int(hashlib.sha256(seed.encode("utf-8")).hexdigest(),16)%3]

def numeric_core(v):
    s=unicodedata.normalize("NFKC",str(v or "")); cn={"一":"1","二":"2","三":"3","四":"4","五":"5","六":"6","七":"7","八":"8","九":"9"}
    for a,b in cn.items(): s=s.replace(a,b)
    m=re.search(r"\d+",s)
    return str(int(m.group())) if m else ""

def load_unit_index():
    wb=openpyxl.load_workbook(ROOT/FILES["unit"],read_only=True,data_only=True); ws=wb[wb.sheetnames[0]]; rows=ws.iter_rows(values_only=True)
    heads=[str(x or "").strip() for x in next(rows)]; ix={h:i for i,h in enumerate(heads)}; index={}
    for r in rows:
        estate=str(r[ix["楼盘"]] or "").strip(); building=str(r[ix["座栋"]] or "").strip(); unit=str(r[ix["单元"]] or "").strip()
        if estate and building: index.setdefault(estate,{}).setdefault(building,set()).add(unit)
    return index

def confirmed_estates(task, selected=None):
    selected=selected or {}
    decisions=dict(task.get("estate_auto",{}))
    for rv in task.get("estate_reviews",[]):
        choice=selected.get(rv["raw"],rv.get("selected",""))
        if choice:
            e=next((c for c in rv["candidates"] if c["楼盘"]==choice),None)
            if e: decisions[rv["raw"]]={"estate":e,"method":"人工确认","confidence":1}
    return decisions

def building_review(task, estate_selected):
    for rv in task.get("estate_reviews",[]): rv["selected"]=estate_selected.get(rv["raw"],"")
    decisions=confirmed_estates(task,estate_selected); index=load_unit_index(); src_by_target={}
    for src,tgts in task["mapping"].items():
        for t in (tgts if isinstance(tgts,list) else [tgts]):
            if t and t!="__ignore__": src_by_target.setdefault(t,[]).append(src)
    community_sources=src_by_target.get("小区",[]); building_sources=src_by_target.get("栋幢",[]); groups={}
    for row in task["rows"]:
        raw_comm=" ".join(str(row.get(s) or "").strip() for s in community_sources).strip(); decision=decisions.get(raw_comm)
        raw_build=" ".join(str(row.get(s) or "").strip() for s in building_sources).strip()
        if decision and raw_build:
            estate=decision["estate"]["楼盘"]; groups.setdefault((estate,raw_build),[]).append(row["_row"])
    reviews=[]
    for (estate,raw),rownums in groups.items():
        buildings=list(index.get(estate,{})); exact=[b for b in buildings if norm(b)==norm(raw)]; core=[b for b in buildings if numeric_core(b) and numeric_core(b)==numeric_core(raw)]
        candidates=exact or core
        if len(candidates)==1: continue
        if not candidates:
            candidates=[b for _,b in sorted(((SequenceMatcher(None,norm(raw,True),norm(b,True)).ratio(),b) for b in buildings),reverse=True)[:8]]
        reviews.append({"key":estate+"|"+raw,"estate":estate,"raw":raw,"rows":rownums,"candidates":candidates,"reason":"多个候选" if len(exact or core)>1 else "无法自动确认","selected":""})
    task["building_reviews"]=reviews
    return reviews

def save_rules(task):
    rules=jload(RULES_FILE,{"mappings":{},"estates":{},"buildings":{},"units":{}})
    rules.setdefault("mappings",{})[task["source_key"]]=task["mapping"]
    for x in task.get("estate_reviews",[]):
        if x.get("selected"): rules.setdefault("estates",{})[x["raw"]]=x["selected"]
    for x in task.get("building_reviews",[]):
        if x.get("selected"): rules.setdefault("buildings",{})[x["key"]]=x["selected"]
    jsave(RULES_FILE,rules)

def build_output(task, clean_only=False):
    kind=task["kind"]; info=template_info(kind); original_sig=task["template_signature"]
    if info["signature"]!=original_sig: raise ValueError("模板在任务开始后发生变化，请新建任务并重新确认字段")
    wb=openpyxl.load_workbook(info["path"]); ws=wb["房源"]
    if ws.max_row>=2: ws.delete_rows(2,ws.max_row-1)
    template_headers=info["headers"]
    headers=template_headers+["未匹配原因"]
    reason_col=len(headers)
    ws.cell(1,reason_col).value="未匹配原因"
    if reason_col>1:
        source_header=ws.cell(1,reason_col-1); reason_header=ws.cell(1,reason_col)
        if source_header.has_style: reason_header._style=copy(source_header._style)
    ws.column_dimensions[openpyxl.utils.get_column_letter(reason_col)].width=36
    src_by_target={}
    for s,ts in task["mapping"].items():
        for t in (ts if isinstance(ts,list) else [ts]):
            if t and t!="__ignore__": src_by_target.setdefault(t,[]).append(s)
    decisions=confirmed_estates(task)
    building_manual={x["key"]:x.get("selected","") for x in task.get("building_reviews",[]) if x.get("selected")}
    audit=[]; exceptions=[]; output=[]; now=datetime.now().isoformat(timespec="seconds"); unit_index=load_unit_index()
    community_sources=src_by_target.get("小区",[])
    for source in task["rows"]:
        record={h:"" for h in headers}
        for target,sources in src_by_target.items():
            if target not in record or SYSTEM_BLANK.search(target): continue
            values=[source.get(s) for s in sources if source.get(s) not in (None,"")]
            if values: record[target]=" ".join(str(v) for v in values)
        raw_comm=" ".join(str(source.get(s) or "").strip() for s in community_sources).strip()
        decision=decisions.get(raw_comm); blocking=[]
        if decision:
            e=decision["estate"]
            for field,dk in (("小区","楼盘"),("城区","行政区"),("商圈","商圈")):
                if field in record and record[field]!=e[dk]:
                    audit.append({"原始行号":source["_row"],"房源编号":record.get("房源编号",""),"字段名称":field,"原始值":record[field],"最终值":e[dk],"使用的规则":decision["method"],"字典来源":FILES["estate"],"匹配依据":raw_comm,"置信度":decision["confidence"],"是否人工确认":decision["method"]=="人工确认","操作时间":now,"规则版本":RULE_VERSION})
                    record[field]=e[dk]
            buildings=unit_index.get(e["楼盘"],{}); raw_build=str(record.get("栋幢") or "").strip(); chosen_build=""
            manual_build=building_manual.get(e["楼盘"]+"|"+raw_build,"")
            exact_build=[b for b in buildings if norm(b)==norm(raw_build)]
            core_build=[b for b in buildings if numeric_core(b) and numeric_core(b)==numeric_core(raw_build)]
            candidates=exact_build or core_build
            if manual_build in buildings: chosen_build=manual_build
            elif raw_build and len(candidates)==1: chosen_build=candidates[0]
            elif raw_build and len(candidates)>1: blocking.append("栋座存在多个冲突候选")
            elif raw_build: blocking.append("栋座无法确认")
            if chosen_build:
                if record.get("栋幢")!=chosen_build:
                    audit.append({"原始行号":source["_row"],"房源编号":record.get("房源编号",""),"字段名称":"栋幢","原始值":record.get("栋幢",""),"最终值":chosen_build,"使用的规则":"栋座人工确认" if manual_build else "楼盘内唯一栋座规范","字典来源":FILES["unit"],"匹配依据":numeric_core(raw_build),"置信度":1,"是否人工确认":bool(manual_build),"操作时间":now,"规则版本":RULE_VERSION})
                    record["栋幢"]=chosen_build
                raw_unit=str(record.get("单元") or "").strip(); units=buildings[chosen_build]
                exact_unit=[u for u in units if norm(u)==norm(raw_unit)]; core_unit=[u for u in units if numeric_core(u) and numeric_core(u)==numeric_core(raw_unit)]; uc=exact_unit or core_unit
                if raw_unit and len(uc)==1:
                    chosen_unit=uc[0]
                    if record.get("单元")!=chosen_unit:
                        audit.append({"原始行号":source["_row"],"房源编号":record.get("房源编号",""),"字段名称":"单元","原始值":record.get("单元",""),"最终值":chosen_unit,"使用的规则":"楼盘栋座内唯一单元规范","字典来源":FILES["unit"],"匹配依据":numeric_core(raw_unit),"置信度":1,"是否人工确认":False,"操作时间":now,"规则版本":RULE_VERSION})
                        record["单元"]=chosen_unit
                elif raw_unit and len(uc)>1: blocking.append("单元存在多个冲突候选")
                elif raw_unit: blocking.append("单元无法确认")
        else: blocking.append("楼盘无法确认")
        price="总价(万)" if kind=="sale" else "月租金"
        if record.get(price) in (None,""): blocking.append(f"完全缺失{price}")
        record["委托日期"]=task["entrust_date"]
        seed=str(record.get("房源编号") or "".join(str(record.get(x) or "") for x in ("小区","栋幢","单元","房号","业主联系方式")))
        record["等级"]=stable_grade(seed)
        for h in headers:
            if SYSTEM_BLANK.search(h): record[h]=""
        record["未匹配原因"]="；".join(dict.fromkeys(blocking))
        if blocking: exceptions.append({"原始行号":source["_row"],"房源编号":record.get("房源编号",""),"原始状态":next((source.get(s) for s in source if "状态" in s),""),"异常":record["未匹配原因"]})
        if not clean_only or not blocking: output.append(record)
    for ri,rec in enumerate(output,2):
        for ci,h in enumerate(headers,1):
            cell=ws.cell(ri,ci); val=rec[h]
            if h=="委托日期":
                try: cell.value=datetime.strptime(val,"%Y-%m-%d").date(); cell.number_format="yyyy-mm-dd"
                except: cell.value=val
            else: cell.value=val
            if TEXT_FIELD.search(h): cell.number_format="@"
            if h=="未匹配原因": cell.alignment=copy(cell.alignment); cell.alignment=cell.alignment.copy(wrap_text=True,vertical="top")
        if ri>2:
            for ci in range(1,len(headers)+1):
                src=ws.cell(2,ci); dst=ws.cell(ri,ci)
                if src.has_style:
                    dst._style=copy(src._style); dst.number_format=src.number_format if not TEXT_FIELD.search(headers[ci-1]) else "@"
    stamp=datetime.now().strftime("%Y%m%d_%H%M%S"); prefix="售房" if kind=="sale" else "租房"
    outdir=OUTPUTS/task["id"]; outdir.mkdir(parents=True,exist_ok=True)
    outpath=outdir/f"{prefix}清洗导入_{stamp}.xlsx"; wb.save(outpath)
    report={"原始文件名":task["original_name"],"类型":"出售" if kind=="sale" else "出租","任务开始时间":task["started"],"委托日期":task["entrust_date"],"原始记录数":len(task["rows"]),"最终输出记录数":len(output),"阻断异常数量":len(exceptions),"规则版本":RULE_VERSION}
    (outdir/"清洗报告.json").write_text(json.dumps(report,ensure_ascii=False,indent=2),"utf-8")
    for name,records in (("审计日志.csv",audit),("异常清单.csv",exceptions)):
        with (outdir/name).open("w",encoding="utf-8-sig",newline="") as f:
            if records:
                w=csv.DictWriter(f,fieldnames=list(records[0])); w.writeheader(); w.writerows(records)
    download_id=register_download(outpath)
    task["last_export"]={"file":str(outpath.relative_to(ROOT)),"downloadId":download_id,"report":report,"audit":len(audit),"exceptions":exceptions}
    return task["last_export"]

class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args): print("[%s] %s"%(self.log_date_time_string(),fmt%args))
    def send_json(self,obj,status=200):
        body=json.dumps(obj,ensure_ascii=False,default=str).encode(); self.send_response(status); self.send_header("Content-Type","application/json; charset=utf-8"); self.send_header("Content-Length",str(len(body))); self.end_headers(); self.wfile.write(body)
    def body(self):
        n=int(self.headers.get("Content-Length",0)); return json.loads(self.rfile.read(n) or b"{}")
    def do_GET(self):
        path=urlparse(self.path).path
        if path=="/api/status": return self.send_json({"files":file_status(),"tasks":len(TASKS)})
        if path.startswith("/download/"):
            token=path[len("/download/"):]
            fp=resolve_download(token)
            if fp is None:
                rel=unquote(token); fp=(ROOT/rel).resolve()
                if ROOT.resolve() not in fp.parents:
                    self.send_response(302); self.send_header("Location","/?download=invalid"); self.end_headers(); return
                if not fp.exists():
                    parent=fp.parent; stamp=re.search(r"(\d{8}_\d{6})\.xlsx$",fp.name,re.I)
                    matches=list(parent.glob(f"*{stamp.group(1)}.xlsx")) if stamp and parent.exists() else []
                    if len(matches)==1: fp=matches[0].resolve()
            if not fp.exists():
                self.send_response(302); self.send_header("Location","/?download=missing"); self.end_headers(); return
            from urllib.parse import quote
            data=fp.read_bytes(); self.send_response(200); self.send_header("Content-Type","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"); self.send_header("Content-Disposition",f"attachment; filename*=UTF-8''{quote(fp.name)}"); self.send_header("Content-Length",str(len(data))); self.end_headers(); return self.wfile.write(data)
        fp=STATIC/("index.html" if path=="/" else path.lstrip("/"))
        if fp.exists():
            data=fp.read_bytes(); mime="text/html" if fp.suffix==".html" else "text/css" if fp.suffix==".css" else "application/javascript"
            self.send_response(200); self.send_header("Content-Type",mime+"; charset=utf-8"); self.send_header("Cache-Control","no-store, max-age=0"); self.send_header("Content-Length",str(len(data))); self.end_headers(); return self.wfile.write(data)
        self.send_error(404)
    def do_POST(self):
        try:
            path=urlparse(self.path).path
            if path=="/api/task":
                b=self.body(); kind=b.get("kind"); missing=[x["name"] for x in file_status() if not x["exists"]]
                if missing: raise ValueError("缺少基础文件："+"、".join(missing))
                if kind not in ("sale","rent"): raise ValueError("请先选择出售或出租")
                info=template_info(kind); tid=uuid.uuid4().hex[:12]; TASKS[tid]={"id":tid,"kind":kind,"started":datetime.now().isoformat(timespec="seconds"),"entrust_date":str(date.today()),"template_signature":info["signature"]}
                return self.send_json({"taskId":tid,"headers":info["headers"]})
            if path=="/api/upload":
                tid=self.headers.get("X-Task-Id"); name=unquote(self.headers.get("X-Filename","source.xlsx")); task=get_task(tid); ext=Path(name).suffix.lower()
                if ext not in (".csv",".xls",".xlsx"): raise ValueError("仅支持 CSV、XLS、XLSX")
                fp=UPLOADS/f"{tid}{ext}"; fp.write_bytes(self.rfile.read(int(self.headers.get("Content-Length",0)))); fp,sheets=inspect_upload(fp)
                task.update({"path":fp,"original_name":name,"sheets":sheets}); return self.send_json({"sheets":sheets,"received":{"name":name,"size":fp.stat().st_size,"sizeMB":round(fp.stat().st_size/1024/1024,1)}})
            if path=="/api/select-sheet":
                b=self.body(); task=get_task(b.get("taskId")); headers,rows=read_table(task["path"],b["sheet"],int(b["headerRow"])); task.update({"sheet":b["sheet"],"header_row":int(b["headerRow"]),"headers":headers,"rows":rows,"source_key":hashlib.sha256("|".join(map(norm,headers)).encode()).hexdigest()})
                target=template_info(task["kind"])["headers"]; return self.send_json({"headers":headers,"rows":len(rows),"suggestions":mapping_suggestions(headers,target,rows,task["source_key"]),"targets":target})
            if path=="/api/mapping":
                b=self.body(); task=get_task(b.get("taskId")); task["mapping"]=b["mapping"]; reviews,errors=estate_review(task); return self.send_json({"reviews":reviews,"errors":errors,"autoCount":len(task.get("estate_auto",{}))})
            if path=="/api/review":
                b=self.body(); task=get_task(b.get("taskId")); selected=b.get("selected",{})
                for rv in task.get("estate_reviews",[]): rv["selected"]=selected.get(rv["raw"],"")
                task["entrust_date"]=b.get("entrustDate",task["entrust_date"]); save_rules(task); return self.send_json({"ok":True})
            if path=="/api/building-review":
                b=self.body(); task=get_task(b.get("taskId")); reviews=building_review(task,b.get("selected",{})); return self.send_json({"reviews":reviews})
            if path=="/api/building-confirm":
                b=self.body(); task=get_task(b.get("taskId")); selected=b.get("selected",{})
                for rv in task.get("building_reviews",[]): rv["selected"]=selected.get(rv["key"],"")
                task["entrust_date"]=b.get("entrustDate",task["entrust_date"]); save_rules(task); return self.send_json({"ok":True})
            if path=="/api/export":
                b=self.body(); return self.send_json(build_output(get_task(b.get("taskId")),bool(b.get("cleanOnly"))))
            self.send_error(404)
        except Exception as e: self.send_json({"error":str(e)},400)

if __name__=="__main__":
    port=int(os.environ.get("HOUSING_CLEANER_PORT","8765")); url=f"http://127.0.0.1:{port}"
    print(f"房源数据清洗工具已启动：{url}")
    threading.Timer(1,lambda:webbrowser.open(url)).start()
    ThreadingHTTPServer(("127.0.0.1",port),Handler).serve_forever()
