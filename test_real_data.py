import json, sys, threading, urllib.request
from urllib.parse import quote
from datetime import date
from pathlib import Path

import openpyxl
import server


def req(url, obj=None, headers=None, raw=None):
    data=raw if raw is not None else (json.dumps(obj,ensure_ascii=False).encode("utf-8") if obj is not None else None)
    hdr=headers or {}
    if obj is not None: hdr["Content-Type"]="application/json"
    request=urllib.request.Request(url,data=data,headers=hdr,method="POST" if data is not None else "GET")
    with urllib.request.urlopen(request,timeout=600) as response: return json.loads(response.read())


def main(path):
    source=Path(path)
    http=server.ThreadingHTTPServer(("127.0.0.1",0),server.Handler)
    threading.Thread(target=http.serve_forever,daemon=True).start()
    base=f"http://127.0.0.1:{http.server_address[1]}"
    try:
        task=req(base+"/api/task",{"kind":"sale"}); tid=task["taskId"]
        upload=req(base+"/api/upload",headers={"X-Task-Id":tid,"X-Filename":quote(source.name)},raw=source.read_bytes())
        sheet=upload["sheets"][0]
        selected=req(base+"/api/select-sheet",{"taskId":tid,"sheet":sheet["name"],"headerRow":sheet["headerRow"]})
        targets=set(selected["targets"])
        mapping={item["source"]:(item["source"] if item["source"] in targets else (item["target"] or "__ignore__")) for item in selected["suggestions"]}
        estate=req(base+"/api/mapping",{"taskId":tid,"mapping":mapping})
        estate_choices={item["raw"]:(item["candidates"][0].get("match_name") or item["candidates"][0]["楼盘"]) for item in estate["reviews"] if item["candidates"]}
        building=req(base+"/api/building-review",{"taskId":tid,"selected":estate_choices})
        building_choices={item["key"]:item["candidates"][0] for item in building["reviews"] if item["candidates"]}
        req(base+"/api/building-confirm",{"taskId":tid,"selected":building_choices,"entrustDate":str(date.today()),"persistRules":False})
        exported=req(base+"/api/export",{"taskId":tid,"cleanOnly":False})
        output_path=server.ROOT/exported["file"]
        workbook=openpyxl.load_workbook(output_path,read_only=True,data_only=False)
        ws=workbook["房源"]
        headers=[str(c.value or "") for c in next(ws.iter_rows(min_row=1,max_row=1))]
        community_col=headers.index("小区")
        building_col=headers.index("栋幢")
        reason_col=headers.index("未匹配原因")
        output_rows=0; reason_rows=0; bracket_phase_rows=0; courtyard_communities=0; leading_apostrophes=0; composite_layout_values=0
        for row in ws.iter_rows(min_row=2,values_only=True):
            if not any(v not in (None,"") for v in row): continue
            output_rows+=1
            community=str(row[community_col] or "")
            if row[reason_col] not in (None,""): reason_rows+=1
            if server.re.search(r"[（(]\s*(?:第?[一二三四五六七八九十\d]+期|[一二三四五六七八九十\d]+区)\s*[）)]",community): bracket_phase_rows+=1
            if server.re.search(r"\d+号院$",community): courtyard_communities+=1
            leading_apostrophes+=sum(isinstance(value,str) and value.startswith("'") for value in row)
            composite_layout_values+=sum(isinstance(row[ix],str) and server.re.search(r"[房室厅厨卫阳台]",row[ix]) is not None for ix in [headers.index(x) for x in ("室","厅","厨","卫","阳台")])
        result={
            "input_file":source.name,
            "reported_rows":sheet["rows"],
            "read_rows":selected["rows"],
            "output_rows":output_rows,
            "estate_auto":estate["autoCount"],
            "estate_reviews":len(estate["reviews"]),
            "building_reviews":len(building["reviews"]),
            "rows_with_reason":reason_rows,
            "phase_parentheses_remaining":bracket_phase_rows,
            "communities_still_ending_courtyard":courtyard_communities,
            "leading_apostrophes_remaining":leading_apostrophes,
            "composite_layout_values_remaining":composite_layout_values,
            "output":str(output_path),
            "report":exported["report"],
        }
        assert sheet["rows"]==selected["rows"]==output_rows
        assert bracket_phase_rows==0
        assert leading_apostrophes==0
        assert composite_layout_values==0
        print(json.dumps(result,ensure_ascii=False))
    finally:
        http.shutdown()


if __name__=="__main__":
    main(sys.argv[1])
