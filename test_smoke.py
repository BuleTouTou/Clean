import json, tempfile, threading, urllib.request
from pathlib import Path
from datetime import date
import openpyxl
import server

def req(url, obj=None, headers=None, raw=None):
    data = raw if raw is not None else (json.dumps(obj).encode() if obj is not None else None)
    h = headers or {}
    if obj is not None: h["Content-Type"] = "application/json"
    r = urllib.request.urlopen(urllib.request.Request(url, data=data, headers=h, method="POST" if data is not None else "GET"))
    return json.loads(r.read())

def main():
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="原始数据"
    ws.append(["北京房源导出表"]); ws.merge_cells("A1:H1")
    ws.append(["楼盘","行政区","板块","栋号","单元","室号","售价","业主电话","录入人工号","维护人工号"])
    ws.append(["连心园","朝阳","常营","1号楼","1单元","101",500,"0013800000000","0260616013","260616013"])
    ws.append(["绝对不存在小区","海淀","中关村","2栋","2","202","", "13900000000","0260616010","260616010"])
    src=server.DATA/"smoke.xlsx"; wb.save(src)
    http=server.ThreadingHTTPServer(("127.0.0.1",0),server.Handler); port=http.server_address[1]
    threading.Thread(target=http.serve_forever,daemon=True).start(); base=f"http://127.0.0.1:{port}"
    task=req(base+"/api/task",{"kind":"sale"}); tid=task["taskId"]
    up=req(base+"/api/upload",headers={"X-Task-Id":tid,"X-Filename":"smoke.xlsx"},raw=src.read_bytes())
    sel=req(base+"/api/select-sheet",{"taskId":tid,"sheet":"原始数据","headerRow":2})
    mapping={x["source"]:(x["target"] or "__ignore__") for x in sel["suggestions"]}
    mapping.update({"楼盘":"小区","行政区":"城区","板块":"商圈","栋号":"栋幢","室号":"房号","售价":"总价(万)","业主电话":"业主联系方式","单元":"单元","录入人工号":"录入人工号","维护人工号":"维护人工号"})
    rev=req(base+"/api/mapping",{"taskId":tid,"mapping":mapping})
    req(base+"/api/review",{"taskId":tid,"selected":{},"entrustDate":str(date.today())})
    out=req(base+"/api/export",{"taskId":tid,"cleanOnly":False}); fp=server.ROOT/out["file"]
    result=openpyxl.load_workbook(fp); ws=result["房源"]
    assert [c.value for c in ws[1]][:-1]==task["headers"] and ws.cell(1,ws.max_column).value=="未匹配原因"
    assert ws.max_row==3 and ws["F2"].value=="连心园" and ws["B2"].is_date
    assert ws["Y2"].value in ("A类","B类","C类") and ws["AB2"].number_format=="@"
    assert str(ws["AL2"].value)=="0260616013" and ws["AL2"].number_format=="@"
    assert str(ws["AP2"].value)=="260616013" and ws["AP2"].number_format=="@"
    assert out["report"]["阻断异常数量"]>=1
    assert any(ws.cell(r,ws.max_column).value for r in range(2,ws.max_row+1))
    print(json.dumps({"ok":True,"output":str(fp),"review":len(rev["reviews"]),"report":out["report"]},ensure_ascii=False))
    http.shutdown()

if __name__=="__main__": main()
